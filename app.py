import os
import uuid
import json
import csv
import io
from datetime import datetime, timedelta, date
from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, Response
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from sqlalchemy import func
from models import db, User, Item, Bill, Sale

app = Flask(__name__)
app.config['SECRET_KEY'] = 'dev-secret-key-123'

# Use environment variable for database URL on production (e.g. PostgreSQL on Vercel)
db_url = os.environ.get('DATABASE_URL') or os.environ.get('POSTGRES_URL')
if db_url and db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url or 'sqlite:///inventory.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

login_manager = LoginManager()
login_manager.login_view = 'login'
login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Initialize database
with app.app_context():
    try:
        db.create_all()
    except Exception as e:
        print(f"Database initialization skipped/failed: {e}")

# --- AUTHENTICATION ROUTES ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        role = request.form.get('role')
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username, role=role).first()
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('index'))
        else:
            flash('Invalid username, password, or role.', 'error')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        role = request.form.get('role') # 'owner' or 'staff'
        username = request.form.get('username')
        password = request.form.get('password')
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'error')
            return redirect(url_for('register'))
            
        new_user = User(
            username=username,
            password_hash=generate_password_hash(password, method='scrypt'),
            role=role
        )
        db.session.add(new_user)
        db.session.commit()
        flash('Registration successful! Please login.', 'success')
        return redirect(url_for('login'))
        
    return render_template('register.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# --- VIEW ROUTES ---

@app.route('/')
@login_required
def index():
    return render_template('index.html', user=current_user)

@app.route('/inventory')
@login_required
def inventory():
    return render_template('inventory.html', user=current_user)

@app.route('/sell')
@login_required
def sell():
    return render_template('sell.html', user=current_user)

@app.route('/invoice')
@login_required
def invoice():
    return render_template('invoice.html', user=current_user)

# --- API ROUTES ---

@app.route('/api/add_item', methods=['POST'])
@login_required
def add_item():
    data = request.json
    name = data.get('name')
    price = data.get('price')
    quantity = data.get('quantity')
    expiry_str = data.get('expiry_date')
    
    if not all([name, price is not None, quantity is not None]):
        return jsonify({'error': 'Missing required fields'}), 400
        
    try:
        price = float(price)
        quantity = int(quantity)
        expiry_date = datetime.strptime(expiry_str, '%Y-%m-%d').date() if expiry_str else None
    except ValueError:
        return jsonify({'error': 'Invalid data formats'}), 400
        
    # Generate item ID
    item_id = f"ITM-{str(uuid.uuid4())[:8].upper()}"
    
    new_item = Item(
        item_id=item_id,
        name=name,
        price=price,
        quantity=quantity,
        expiry_date=expiry_date
    )
    db.session.add(new_item)
    db.session.commit()
    
    return jsonify({
        'message': 'Item added successfully',
        'item': {
            'item_id': item_id,
            'name': name,
            'price': price,
            'quantity': quantity,
            'expiry_date': str(expiry_date) if expiry_date else None
        }
    }), 201

@app.route('/api/edit_item/<item_id>', methods=['POST'])
@login_required
def edit_item(item_id):
    data = request.json
    item = Item.query.filter_by(item_id=item_id).first()
    if not item:
        return jsonify({'error': 'Item not found'}), 404
        
    try:
        if 'name' in data:
            item.name = data['name']
        if 'price' in data:
            item.price = float(data['price'])
        if 'quantity' in data:
            item.quantity = int(data['quantity'])
        if 'expiry_date' in data:
            item.expiry_date = datetime.strptime(data['expiry_date'], '%Y-%m-%d').date() if data['expiry_date'] else None
    except ValueError:
        return jsonify({'error': 'Invalid data formats'}), 400
        
    db.session.commit()
    return jsonify({'message': 'Item updated successfully'})

@app.route('/api/inventory', methods=['GET'])
@login_required
def get_inventory():
    items = Item.query.all()
    inventory_data = []
    for item in items:
        inventory_data.append({
            'item_id': item.item_id,
            'name': item.name,
            'price': item.price,
            'quantity': item.quantity,
            'expiry_date': item.expiry_date.strftime('%Y-%m-%d') if item.expiry_date else None,
            'added_date': item.added_date.strftime('%Y-%m-%d %H:%M:%S')
        })
    return jsonify({'inventory': inventory_data})

@app.route('/api/sell', methods=['POST'])
@login_required
def api_sell():
    data = request.json
    cart = data.get('cart', [])
    discount_pct = float(data.get('discount_pct', 0))
    
    if not cart:
        return jsonify({'error': 'Cart is empty'}), 400
        
    subtotal = 0
    sales_to_record = []
    
    # Validation and Stock Check
    for item_data in cart:
        qty = int(item_data.get('quantity', 0))
        item_id_str = item_data.get('item_id')
        
        if qty <= 0:
            return jsonify({'error': f'Invalid quantity for item {item_id_str}'}), 400
            
        item = Item.query.filter_by(item_id=item_id_str).first()
        if not item:
            return jsonify({'error': f'Item {item_id_str} not found'}), 404
            
        if item.quantity < qty:
            return jsonify({'error': f'Insufficient stock for {item.name}'}), 400
            
        item_subtotal = item.price * qty
        subtotal += item_subtotal
        
        sales_to_record.append({
            'item': item,
            'qty': qty,
            'item_subtotal': item_subtotal
        })

    # Calculations
    discount_amount = subtotal * (discount_pct / 100)
    after_discount = subtotal - discount_amount
    gst_amount = after_discount * 0.18
    total_amount = after_discount + gst_amount
    
    bill_id = f"BILL-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
    
    # Record Bill
    new_bill = Bill(
        bill_id=bill_id,
        subtotal=subtotal,
        discount_amount=discount_amount,
        gst_amount=gst_amount,
        total_amount=total_amount
    )
    db.session.add(new_bill)
    db.session.flush() # To get new_bill.id
    
    # Record Sales & Update Stock
    receipt_items = []
    for sale_data in sales_to_record:
        item = sale_data['item']
        qty = sale_data['qty']
        
        item.quantity -= qty
        
        new_sale = Sale(
            bill_id=new_bill.id,
            item_id=item.id,
            quantity=qty
        )
        db.session.add(new_sale)
        
        receipt_items.append({
            'name': item.name,
            'quantity': qty,
            'price': item.price,
            'total': sale_data['item_subtotal']
        })
        
    db.session.commit()
    
    return jsonify({
        'message': 'Sale successful',
        'bill': {
            'bill_id': bill_id,
            'date': new_bill.date.strftime('%Y-%m-%d %H:%M:%S'),
            'items': receipt_items,
            'subtotal': round(subtotal, 2),
            'discount': round(discount_amount, 2),
            'gst': round(gst_amount, 2),
            'total': round(total_amount, 2)
        }
    })

@app.route('/api/alerts', methods=['GET'])
@login_required
def get_alerts():
    items = Item.query.all()
    today = date.today()
    alerts = []
    
    for item in items:
        # Stock Alerts
        if item.quantity == 0:
            alerts.append({'type': 'OUT_OF_STOCK', 'message': f'{item.name} is out of stock!', 'severity': 'danger'})
        elif item.quantity < 5:
            alerts.append({'type': 'LOW_STOCK', 'message': f'{item.name} has low stock ({item.quantity} left).', 'severity': 'warning'})
            
        # Expiry Alerts
        if item.expiry_date:
            delta = (item.expiry_date - today).days
            if delta < 0:
                alerts.append({'type': 'EXPIRED', 'message': f'{item.name} expired on {item.expiry_date}!', 'severity': 'danger'})
            elif delta < 30:
                alerts.append({'type': 'NEAR_EXPIRY', 'message': f'{item.name} is expiring very soon ({delta} days).', 'severity': 'warning'})
            elif delta < 90:
                alerts.append({'type': 'EXPIRING_SOON', 'message': f'{item.name} expires in {delta} days.', 'severity': 'info'})
                
    return jsonify({'alerts': alerts})

@app.route('/api/dashboard_data', methods=['GET'])
@login_required
def dashboard_data():
    # 1. Product sold (Bar chart)
    # 2. Daily revenue (Line chart)
    
    # Group by item id for total quantity sold
    sales_grouped = db.session.query(
        Item.name, 
        func.sum(Sale.quantity).label('total_qty')
    ).join(Sale).group_by(Item.id).all()
    
    products_sold = sorted([{'name': row[0], 'qty': row[1]} for row in sales_grouped], key=lambda x: x['qty'], reverse=True)[:10]
    top_product = products_sold[0]['name'] if products_sold else 'None'
    
    # Revenue grouped by date
    # Format date as YYYY-MM-DD for grouping
    bills = db.session.query(
        db.cast(Bill.date, db.Date).label('b_date'),
        func.sum(Bill.total_amount).label('daily_total')
    ).group_by(db.cast(Bill.date, db.Date)).order_by(db.cast(Bill.date, db.Date).desc()).limit(14).all()
    
    daily_revenue = [{'date': str(row[0]), 'total': float(row[1])} for row in reversed(bills)]
    
    # Sales Velocity
    today = datetime.utcnow()
    velocity_data = []
    
    items = Item.query.all()
    for item in items:
        total_sold = db.session.query(func.sum(Sale.quantity)).filter_by(item_id=item.id).scalar() or 0
        if total_sold > 0:
            days_active = max(1, (today - item.added_date).days)
            velocity = round(total_sold / days_active, 2)
            velocity_data.append({'name': item.name, 'velocity': velocity})
            
    velocity_data = sorted(velocity_data, key=lambda x: x['velocity'], reverse=True)[:10]
    
    return jsonify({
        'products_sold': products_sold,
        'daily_revenue': daily_revenue,
        'velocity_data': velocity_data,
        'top_product': top_product
    })

@app.route('/api/download_template')
@login_required
def download_template():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Medicine Name', 'Price (INR)', 'Stock Quantity', 'Expiry Date (YYYY-MM-DD)'])
    writer.writerow(['Paracetamol 650mg', '15.50', '100', '2027-12-31'])
    writer.writerow(['Amoxicillin 500mg', '85.00', '50', '2026-10-15'])
    writer.writerow(['Cetirizine 10mg', '5.20', '200', ''])
    
    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=inventory_template.csv'
    return response

@app.route('/api/upload_inventory', methods=['POST'])
@login_required
def upload_inventory():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
        
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    filename = file.filename.lower()
    rows = []
    
    try:
        if filename.endswith('.csv'):
            file_content = file.read().decode('utf-8-sig')
            csv_data = csv.DictReader(io.StringIO(file_content))
            for r in csv_data:
                rows.append(r)
        elif filename.endswith(('.xlsx', '.xls')):
            import openpyxl
            file_content = file.read()
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            if not any(headers):
                return jsonify({'error': 'Empty Excel sheet'}), 400
                
            for row_idx in range(2, sheet.max_row + 1):
                row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, len(headers) + 1)]
                if not any(v is not None for v in row_values):
                    continue
                row_dict = {headers[i]: row_values[i] for i in range(len(headers)) if headers[i] is not None}
                rows.append(row_dict)
        else:
            return jsonify({'error': 'Unsupported file format. Please upload CSV or Excel (.xlsx/.xls)'}), 400
    except Exception as e:
        return jsonify({'error': f'Failed to read file: {str(e)}'}), 400

    if not rows:
        return jsonify({'error': 'No data rows found in file'}), 400

    added_count = 0
    updated_count = 0
    errors = []

    def get_field(row_dict, options):
        for opt in options:
            for k in row_dict.keys():
                if k and k.strip().lower() == opt.lower():
                    return row_dict[k]
        return None

    for idx, row in enumerate(rows, start=2):
        name_val = get_field(row, ['medicine name', 'item name', 'name', 'product name'])
        price_val = get_field(row, ['price (inr)', 'price', 'list price', 'mrp'])
        qty_val = get_field(row, ['stock quantity', 'quantity', 'stock qty', 'stock', 'qty'])
        expiry_val = get_field(row, ['expiry date', 'expiry date (yyyy-mm-dd)', 'expiry'])

        if not name_val or str(name_val).strip() == '':
            errors.append({'row': idx, 'error': 'Missing medicine/item name'})
            continue
            
        name_val = str(name_val).strip()

        if price_val is None or str(price_val).strip() == '':
            errors.append({'row': idx, 'error': f"Missing price for '{name_val}'"})
            continue
        try:
            price = float(str(price_val).replace('₹', '').replace(',', '').strip())
            if price < 0:
                errors.append({'row': idx, 'error': f"Price cannot be negative for '{name_val}'"})
                continue
        except ValueError:
            errors.append({'row': idx, 'error': f"Invalid price format '{price_val}' for '{name_val}'"})
            continue

        if qty_val is None or str(qty_val).strip() == '':
            errors.append({'row': idx, 'error': f"Missing quantity for '{name_val}'"})
            continue
        try:
            quantity = int(float(str(qty_val).replace(',', '').strip()))
            if quantity < 0:
                errors.append({'row': idx, 'error': f"Quantity cannot be negative for '{name_val}'"})
                continue
        except ValueError:
            errors.append({'row': idx, 'error': f"Invalid quantity format '{qty_val}' for '{name_val}'"})
            continue

        expiry_date = None
        if expiry_val and str(expiry_val).strip() != '':
            expiry_str = str(expiry_val).strip()
            parsed = False
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d', '%d/%m/%Y'):
                try:
                    if ' ' in expiry_str:
                        expiry_str = expiry_str.split(' ')[0]
                    expiry_date = datetime.strptime(expiry_str, fmt).date()
                    parsed = True
                    break
                except ValueError:
                    continue
            
            if not parsed and isinstance(expiry_val, (datetime, date)):
                expiry_date = expiry_val if isinstance(expiry_val, date) else expiry_val.date()
                parsed = True
                
            if not parsed:
                errors.append({'row': idx, 'error': f"Invalid expiry date format '{expiry_val}' for '{name_val}'. Use YYYY-MM-DD."})
                continue

        try:
            existing_item = Item.query.filter(func.lower(Item.name) == func.lower(name_val)).first()
            if existing_item:
                existing_item.price = price
                existing_item.quantity += quantity
                if expiry_date:
                    existing_item.expiry_date = expiry_date
                updated_count += 1
            else:
                item_id = f"ITM-{str(uuid.uuid4())[:8].upper()}"
                new_item = Item(
                    item_id=item_id,
                    name=name_val,
                    price=price,
                    quantity=quantity,
                    expiry_date=expiry_date
                )
                db.session.add(new_item)
                added_count += 1
        except Exception as db_err:
            errors.append({'row': idx, 'error': f"Database save error: {str(db_err)}"})

    try:
        db.session.commit()
    except Exception as commit_err:
        db.session.rollback()
        return jsonify({'error': f"Failed to save changes: {str(commit_err)}"}), 500

    return jsonify({
        'success': True,
        'added': added_count,
        'updated': updated_count,
        'errors': errors
    })

if __name__ == '__main__':
    app.run(debug=True, port=5000)
